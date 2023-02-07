# -*- coding: utf-8 -*-
"""
Created on Sat Jan  7 14:45:06 2023

@author: steven
"""
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
import random
import datetime
import time
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment

def RandomTimeSleep():
    RandomNum = random.uniform(5, 10)
    time.sleep(RandomNum)

result=[]
ids = []
data = {}

now = datetime.datetime.now()

hotellist = ['h486567', # 金典
             'h1600402', # 福華
             'h4612157', # 裕元
             'h16411420', # 全國
             'h5442428', # 福容
             'h1340364', # 長榮
             'h8814584', # 清新
             'h10027150', # 林酒店
             'h5494391', # 日月千禧
             'h24890206', # 大毅老爺
             'h8308311', # 日光溫泉
             #'le-meridien-taichung', # 台中艾美
             'h89799999', # 震大金鬱金香酒店
             'h12614734', # 兆品酒店 - 兆尹樓
             'h996969' # 兆品酒店 - 品臻樓
             ]

checkin = now.strftime("%Y-%m-%d")
checkout = (now + datetime.timedelta(days=1)).strftime("%Y-%m-%d")

excel_filename = './OTAExcel/Expedia.xlsx'

for hotelnameen in tqdm(hotellist):
    RandomTimeSleep()
    
    url = 'https://www.expedia.com.tw/'+hotelnameen+'.Hotel-Information?chkin='+checkin+'&chkout='+checkout+'&x_pwa=1&rfrr=HSR&pwa_ts=1673073831616&referrerUrl=aHR0cHM6Ly93d3cuZXhwZWRpYS5jb20udHcvSG90ZWwtU2VhcmNo&useRewards=false&rm1=a1&regionId=3586&destination=%E5%8F%B0%E4%B8%AD%2C+%E5%8F%B0%E7%81%A3&destType=CURRENT_LOCATION&neighborhoodId=6177566&selected=1600402&sort=RECOMMENDED&top_dp=5368&top_cur=TWD&userIntent=&selectedRoomType=200042026&selectedRatePlan=200242690'
    
    chrome_options = Options()
    
    chrome_options.add_argument('User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('blink-setting=imagesEnabled=false')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-blink-features')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--headless')
    
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    
    driver = webdriver.Chrome(options=chrome_options)
    
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source" : "Object.defineProperty(navigator, 'webdriver', {get: () => underfined})"})
    
    driver.get(url)
    RandomTimeSleep()
        
    driver.find_elements(By.CLASS_NAME, "uitk-layout-grid uitk-layout-grid-has-columns-by-auto_fill uitk-layout-grid-has-columns-using-auto-grid uitk-layout-grid-has-space uitk-layout-grid-display-grid uitk-layout-grid-justify-content-start")
    
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    
    driver.close()
    driver.quit()
    
    try:
        div = soup.find_all("div")
    except:
        div = None
    
    ids = []
    for i in range(0,len(div)):
        try:
            id = div[i].get('data-stid')
        except:
            id = None
            
        if(id is not None and 'property-offer-' in id):
            ids.append(id)

    if len(ids) == 0:
        data["RoomName"] = '您選擇的日期在 Expedia 上已無空房，試試其他日期以查看供應情況。'
        data["Original"] = None
        data["Price"] = None
        data["HotelName"] = soup.find("h1",{"class":"uitk-heading uitk-heading-4"}).text.replace("\n","")
        result.append(data)
        data = {}
    else:
        for idx in range(0,len(ids)):
            allData = soup.find("div",{"data-stid":ids[idx]})
                    
            data["RoomName"] = allData.find("h3",{"class":"uitk-heading uitk-heading-6"}).text.replace("\n","").strip()
                           
            try:
                data["Original"] = allData.find("div",{"uitk-text uitk-type-start uitk-type-200 uitk-text-default-theme"}).text.replace("\n","").strip().strip('總價 NT$\xa0').replace(',','')
            except:
                data["Original"] = allData.find("div",{"uitk-text uitk-type-300 uitk-type-bold uitk-text-negative-theme"}).text.replace("\n","") 
            
            try:
                data["Price"] = allData.find("div",{"uitk-text uitk-type-600 uitk-type-bold uitk-text-emphasis-theme"}).text.replace("\n","").strip().strip('NT$\xa0').replace(',','')
            except:
                data["Price"] = allData.find("div",{"uitk-text uitk-type-300 uitk-type-bold uitk-text-negative-theme"}).text.replace("\n","") 
         
            data["HotelName"] = soup.find("h1",{"class":"uitk-heading uitk-heading-4"}).text.replace("\n","") 
            
            if(data is not None):
                result.append(data)
            data = {}

df = pd.DataFrame(result, columns=['HotelName', 'RoomName', 'Original', 'Price'])
df = df.rename(columns={'HotelName': '飯店名稱', 'RoomName': '房型', 'Original': '原價', 'Price': '售價'})
df.head()

df.to_excel(excel_filename, index=None, startrow = 1, sheet_name = "Expedia")

wb = openpyxl.load_workbook(excel_filename)
ws = wb.active
ws['A1'].value = '訂房日期'+checkin+' - '+checkout
align = Alignment(horizontal='center', vertical='center',wrap_text=True)

ws['A1'].alignment = align

ws.merge_cells('A1:D1')

wb.save(excel_filename)


